import openpyxl
import datetime
import pickle
import os
import re
import xlwt
import copy

from suds.client import Client
from suds.sax.element import Element

from constFunction import function_log, manual_error_log
from SqlFunction import create_session, generate_key
from SqlFunction import tVoucher, tEntry, tBaseKdAccount, tBaseKdEntity, tBaseKdSupplier, tBaseKdCostCenter, \
    tWorkOrderMain, tBaseClosingArgs, tBaseClosingEntityInfo, tBaseClosingSmallTaxeeArgs, tBackupTBObject

Domain = '#'


class TBLine:
    def __init__(self, tb_line: list):
        self.balance_type = tb_line[0]
        self.entity_code = tb_line[1]
        self.entity_name = tb_line[2]
        self.month = tb_line[3]
        self.year = tb_line[4]
        self.account_code = tb_line[5]
        self.account_name = tb_line[6]
        self.currency_code = tb_line[7]
        self.currency_name = tb_line[8]
        self.begin_balance_ori_currency = tb_line[9]
        self.year_debit_ori_currency = tb_line[10]
        self.year_credit_ori_currency = tb_line[11]
        self.period_debit_ori_currency = tb_line[12]
        self.period_credit_ori_currency = tb_line[13]
        self.end_balance_ori_currency = tb_line[14]
        self.begin_balance_local_currency = tb_line[15]
        self.year_debit_local_currency = tb_line[16]
        self.year_credit_local_currency = tb_line[17]
        self.period_debit_local_currency = tb_line[18]
        self.period_credit_local_currency = tb_line[19]
        self.end_balance_local_currency = tb_line[20]
        self.begin_balance_report_currency = tb_line[21]
        self.year_debit_report_currency = tb_line[22]
        self.year_credit_report_currency = tb_line[23]
        self.period_debit_report_currency = tb_line[24]
        self.period_credit_report_currency = tb_line[25]
        self.end_balance_report_currency = tb_line[26]
        self.year_pl_ori_currency = tb_line[27]
        self.period_pl_ori_currency = tb_line[28]
        self.year_pl_local_currency = tb_line[29]
        self.period_pl_local_currency = tb_line[30]
        self.year_pl_report_currency = tb_line[31]
        self.period_pl_report_currency = tb_line[32]
        self.begin_qty = tb_line[33]
        self.year_debit_qty = tb_line[34]
        self.year_credit_qty = tb_line[35]
        self.period_debit_qty = tb_line[36]
        self.period_credit_qty = tb_line[37]
        self.end_qty = tb_line[38]
        self.asst_code = tb_line[39]
        self.asst_name = tb_line[40]
        self.has_asst = tb_line[41]


class AsstTBLine:
    def __init__(self, tb_line: list):
        self.balance_type = tb_line[0]
        self.entity_code = tb_line[1]
        self.entity_name = tb_line[2]
        self.month = tb_line[3]
        self.year = tb_line[4]
        self.account_code = tb_line[5]
        self.account_name = tb_line[6]
        self.currency_code = tb_line[7]
        self.currency_name = tb_line[8]
        self.begin_balance_ori_currency = tb_line[9]
        self.year_debit_ori_currency = tb_line[10]
        self.year_credit_ori_currency = tb_line[11]
        self.period_debit_ori_currency = tb_line[12]
        self.period_credit_ori_currency = tb_line[13]
        self.end_balance_ori_currency = tb_line[14]
        self.begin_balance_local_currency = tb_line[15]
        self.year_debit_local_currency = tb_line[16]
        self.year_credit_local_currency = tb_line[17]
        self.period_debit_local_currency = tb_line[18]
        self.period_credit_local_currency = tb_line[19]
        self.end_balance_local_currency = tb_line[20]
        self.begin_balance_report_currency = tb_line[21]
        self.year_debit_report_currency = tb_line[22]
        self.year_credit_report_currency = tb_line[23]
        self.period_debit_report_currency = tb_line[24]
        self.period_credit_report_currency = tb_line[25]
        self.end_balance_report_currency = tb_line[26]
        self.year_pl_ori_currency = tb_line[27]
        self.period_pl_ori_currency = tb_line[28]
        self.year_pl_local_currency = tb_line[29]
        self.period_pl_local_currency = tb_line[30]
        self.year_pl_report_currency = tb_line[31]
        self.period_pl_report_currency = tb_line[32]
        self.begin_qty = tb_line[33]
        self.year_debit_qty = tb_line[34]
        self.year_credit_qty = tb_line[35]
        self.period_debit_qty = tb_line[36]
        self.period_credit_qty = tb_line[37]
        self.end_qty = tb_line[38]
        self.asst_code = tb_line[39]
        self.asst_name = tb_line[40]
        self.asst_full_desc = tb_line[41]
        self.asst_full_name = tb_line[42]
        self.asst_full_code = tb_line[43]
        # parse asst information
        asst_type_list = self.asst_full_code.split(';')
        self.asst_type1 = asst_type_list[0].split('_!')[1].split(':')[0]
        self.asst_code1 = asst_type_list[0].split('_!')[1].split(':')[1]
        try:
            self.asst_type2 = asst_type_list[1].split('_!')[1].split(':')[0]
            self.asst_code2 = asst_type_list[1].split('_!')[1].split(':')[1]
        except IndexError:
            self.asst_type2 = None
            self.asst_code2 = None


class AsstTB:
    def __init__(self, tb: list):
        self.asst_list = []
        if tb:
            for each in tb:
                if not re.match('^\\d', str(each[0])):
                    continue
                self.asst_list.append(AsstTBLine(each))


class TB:
    def __init__(self, tb: list):
        """
        tested the size of 2021-12 hq tb, with all asst accounts info, the size was approximately 24.4M
        which should be considered acceptable
        :param tb:
        """
        self.record_list = []
        self.asst_tb_dict = {}
        self._small_taxee_total_revenue = None
        self._small_taxee_season_rev_dict = {}
        if tb:
            for line in tb:
                if not re.match('^\\d', str(line[0])):
                    continue
                self.record_list.append(TBLine(line))

    def get_asst_tb(self, account_code) -> AsstTB:
        if not self.asst_tb_dict.get(account_code):
            ssid = login_service()
            self.asst_tb_dict[account_code] = AsstTB(get_assit_balance(self.record_list[0].entity_code, account_code,
                                                                       self.record_list[0].year,
                                                                       self.record_list[0].month, ssid))
        return self.asst_tb_dict[account_code]

    def get_all_asst_tb(self) -> None:
        for each in self.record_list:
            if self.is_leaf_account(each.account_code) and each.has_asst == 0:
                self.get_asst_tb(each.account_code)

    def is_leaf_account(self, account_code) -> bool:
        for each in self.record_list:
            if re.match(account_code + '.+', each.account_code):
                return False
        return True

    def _get_previous_tbs(self, month_step_list: list, allow_online_update=False) -> list:
        pre_tb_period_list = []
        for step in month_step_list:
            target_year = int(self.record_list[0].year)
            target_month = int(self.record_list[0].month) + step
            while target_month < 1 or target_month > 12:
                if target_month < 1:
                    target_month += 12
                    target_year -= 1
                elif target_month > 12:
                    target_month -= 12
                    target_year += 1
            pre_tb_period_list.append((target_year, target_month))
        pre_session = create_session()
        entity_id = pre_session.query(tBaseKdEntity.r_id).filter(
            tBaseKdEntity.r_code == self.record_list[0].entity_code
        ).first()[0]
        result_obj_list = []
        for period in pre_tb_period_list:
            fp = pre_session.query(tBackupTBObject.r_file_storage_full_path).filter(
                tBackupTBObject.r_entity_id == entity_id,
                tBackupTBObject.r_year == int(period[0]),
                tBackupTBObject.r_month == int(period[1])
            ).first()[0]
            if fp:
                try:
                    with open(fp, 'rb') as p:
                        tb = pickle.load(p)
                except Exception:
                    tb = None
            elif allow_online_update:
                tb = TB(get_account_balance(
                    entity_code=self.record_list[0].entity_code,
                    str_year=str(period[0]),
                    str_month=str(period[1]),
                    ssid=login_service()
                ))
                tb.get_all_asst_tb()
            else:
                tb = None
            result_obj_list.append(tb)
        return result_obj_list

    def get_small_taxee_revenue_total(self, allow_online_update=False):
        if self._small_taxee_total_revenue:
            return self._small_taxee_total_revenue
        self_session = create_session()
        entity_is_small = self_session.query(tBaseClosingEntityInfo.r_is_small).filter(
            tBaseClosingEntityInfo.r_id == tBaseKdEntity.r_id,
            tBaseKdEntity.r_code == self.record_list[0].entity_code
        ).first()[0]
        if entity_is_small != 0 or int(self.record_list[0].month) not in (3, 6, 9, 12):
            return None
        tb_obj_list = [self, self._get_previous_tbs(month_step_list=[-1, -2], allow_online_update=allow_online_update)]
        rev_acc_record = self_session.query(tBaseClosingSmallTaxeeArgs).filter(
            tBaseClosingSmallTaxeeArgs.r_avaiable == 0,
            tBaseClosingSmallTaxeeArgs.r_is_cal_base_acc == 0
        ).all()
        gen_restoration_rate = float(self_session.query(tBaseClosingArgs).filter(
            tBaseClosingArgs.r_code == 'smallVATRestorationRate'
        ).first()[0])
        rev_acc_dict = {}
        for record in rev_acc_record:
            rev_acc_dict[record.r_account_code] = float(record.r_restoration_rate)
        for tb in tb_obj_list:
            if tb is None:
                return None
        self._small_taxee_total_revenue = 0
        for tb in tb_obj_list:
            for bal in tb.record_list:
                if (bal.balance_type == '1'
                        and bal.currency_code == 'GLC'
                        and bal.account_code in rev_acc_dict):
                    if bal.account_code not in self._small_taxee_season_rev_dict:
                        self._small_taxee_season_rev_dict[bal.account_code] = {}
                    if bal.has_asst == '0':
                        if 'None' not in self._small_taxee_season_rev_dict[bal.account_code]:
                            self._small_taxee_season_rev_dict[bal.account_code]['None'] = 0
                        self._small_taxee_season_rev_dict[bal.account_code]['None'] += float(bal.period_pl_local_currency)
                    else:
                        asst_record_obj = tb.get_asst_tb(bal.account_code)
                        for asst_record in asst_record_obj.asst_list:
                            if (asst_record.balance_type == '1'
                                    and asst_record.currency_code == 'GLC'):
                                if asst_record.asst_code1 not in self._small_taxee_season_rev_dict[bal.account_code]:
                                    self._small_taxee_season_rev_dict[bal.account_code][asst_record.asst_code1] = 0
                                self._small_taxee_season_rev_dict[bal.account_code][asst_record.asst_code1] += float(
                                    asst_record.period_pl_local_currency
                                )
                    self._small_taxee_total_revenue += ((float(bal.period_pl_local_currency)
                                                         * (1 + rev_acc_dict[bal.account_code])))
        self._small_taxee_total_revenue = self._small_taxee_total_revenue / (1 + gen_restoration_rate)
        return self._small_taxee_total_revenue

    def get_small_taxee_revenue_record(self, account_code: str) -> dict:
        """
        {account_code: {asst_code: amount}}
        :param account_code:
        :return: {asst_code: amount}
        """
        if not self._small_taxee_season_rev_dict:
            self.get_small_taxee_revenue_total()
        return self._small_taxee_season_rev_dict.get(account_code)

    def guess_asst_type(self, asst_code: str) -> str:
        type_str = '成本中心' if re.match('^\\d+$', asst_code) else '客户'
        return type_str


def _get_standard_voucher(assit=False):
    standard_voucher = {
        'companyNumber': '',
        'bookedDate': '',
        'bizDate': '',
        'periodYear': 0,
        'periodNumber': 0,
        'voucherType': '',
        'voucherNumber': '',
        'entrySeq': 0,
        'voucherAbstract': '',
        'accountNumber': '',
        'currencyNumber': '',
        'localRate': 1,
        'entryDC': 0,
        'originalAmount': 0.0,
        'debitAmount': 0.0,
        'creditAmount': 0.0,
        'creator': '',
    }
    standard_assit_voucher = {
        'companyNumber': '',
        'bookedDate': '',
        'bizDate': '',
        'periodYear': 0,
        'periodNumber': 0,
        'voucherType': '',
        'voucherNumber': '',
        'entrySeq': 0,
        'voucherAbstract': '',
        'accountNumber': '',
        'currencyNumber': '',
        'localRate': 1,
        'entryDC': 0,
        'originalAmount': 0.0,
        'debitAmount': 0.0,
        'creditAmount': 0.0,
        'creator': '',
        'asstActType1': '',
        'asstActNumber1': '',
    }
    if assit:
        return standard_assit_voucher
    else:
        return standard_voucher


def _importVoucher(voucher_list: list, ssid: str) -> dict:
    """
    this function only invokes Kingdea network interface
    :param voucher_list:
    :param ssid:
    :return: kingdae returns a a list, which is in form of ["error code||voucher type||year||month||chs result||voucher number||result code"]
            the function result should be in form of :
            {
                'result': str,
                'type': str,
                'year': str,
                'month: str,
                'error_msg': str
                'voucher_number': str,
                'origin_result_str': str,
                'desc': str,
            }
    """
    global Domain
    # 开始引入
    try:
        client = Client(Domain + '/ormrpc/services/WSGLWebServiceFacade?wsdl')
        client.set_options(timeout=300)
        ssn = Element('SessionId').setText(ssid).setPrefix(p='ns1', u='http://login.webservice.bos.kingdee.com')
        client.set_options(soapheaders=ssn)
        result = client.service.importVoucher(voucher_list, 1, 0, 0)
        parse_result = result[0].split('||')
        return {
            'result': parse_result[0],
            'type': parse_result[1],
            'year': parse_result[2],
            'month': parse_result[3],
            'invoke_msg': parse_result[4],
            'voucher_number': parse_result[5],
            'origin_result_str': result[0],
        }
    except Exception as e:
        manual_error_log(str(e))
        return {
            'result': '-2',
            'type': None,
            'year': None,
            'month': None,
            'invoke_msg': None,
            'voucher_number': None,
            'origin_result_str': str(e),
        }


def _seperate_voucher_list(voucher_list: list, renumber=False) -> dict:
    """
    clean the voucher list
    :param voucher_list: use voucher number and company number as unique key to identify voucher,
     the key to result dictionary is 'voucherNumber_companyNumber'
    :param renumber: WARNING! if this args is True,
     the function will comb the list and renumber the sequence of entry, which may lead unknown error.
    :return: the dict of voucher list, the key should be the voucher number
    """
    seperated_voucher_dict = {}
    for voucher in voucher_list:
        voucher_key = '_'.join([str(voucher['voucherNumber']), str(voucher['companyNumber'])])
        if voucher_key not in seperated_voucher_dict:
            seperated_voucher_dict.setdefault(voucher_key, [])
        seperated_voucher_dict[voucher_key].append(voucher)
    if renumber:
        for voucher_key in seperated_voucher_dict:
            for new_entry_int, voucher in enumerate(seperated_voucher_dict[voucher_key]):
                voucher['entrySeq'] = new_entry_int + 1
    return seperated_voucher_dict


def export_to_excel(voucher_list: list, file_name_without_tail: str, module='openpyxl'):
    if module == 'openpyxl':
        start_count = 1
        wb = openpyxl.Workbook()
        sht = wb.create_sheet('凭证')
        sht_cf = wb.create_sheet('现金流量')
        sht.cell(start_count, start_count + 0, '公司')
        sht.cell(start_count, start_count + 1, '记账日期')
        sht.cell(start_count, start_count + 2, '业务日期')
        sht.cell(start_count, start_count + 3, '会计期间')
        sht.cell(start_count, start_count + 4, '凭证类型')
        sht.cell(start_count, start_count + 5, '凭证号')
        sht.cell(start_count, start_count + 6, '分录号')
        sht.cell(start_count, start_count + 7, '摘要')
        sht.cell(start_count, start_count + 8, '科目')
        sht.cell(start_count, start_count + 9, '科目名称')
        sht.cell(start_count, start_count + 10, '币种')
        sht.cell(start_count, start_count + 11, '汇率')
        sht.cell(start_count, start_count + 12, '方向')
        sht.cell(start_count, start_count + 13, '原币金额')
        sht.cell(start_count, start_count + 14, '数量')
        sht.cell(start_count, start_count + 15, '单价')
        sht.cell(start_count, start_count + 16, '借方金额')
        sht.cell(start_count, start_count + 17, '贷方金额')
        sht.cell(start_count, start_count + 18, '制单人')
        sht.cell(start_count, start_count + 19, '过账人')
        sht.cell(start_count, start_count + 20, '审核人')
        sht.cell(start_count, start_count + 21, '附件数量')
        sht.cell(start_count, start_count + 22, '过账标记')
        sht.cell(start_count, start_count + 23, '机制凭证模块')
        sht.cell(start_count, start_count + 24, '删除标记')
        sht.cell(start_count, start_count + 25, '凭证序号')
        sht.cell(start_count, start_count + 26, '单位')
        sht.cell(start_count, start_count + 27, '参考信息')
        sht.cell(start_count, start_count + 28, '是否有现金流量')
        sht.cell(start_count, start_count + 29, '现金流量标记')
        sht.cell(start_count, start_count + 30, '业务编号')
        sht.cell(start_count, start_count + 31, '结算方式')
        sht.cell(start_count, start_count + 32, '结算号')
        sht.cell(start_count, start_count + 33, '辅助账摘要')
        sht.cell(start_count, start_count + 34, '核算项目1')
        sht.cell(start_count, start_count + 35, '编码1')
        sht.cell(start_count, start_count + 36, '名称1')
        sht.cell(start_count, start_count + 37, '核算项目2')
        sht.cell(start_count, start_count + 38, '编码2')
        sht.cell(start_count, start_count + 39, '名称2')
        sht.cell(start_count, start_count + 40, '核算项目3')
        sht.cell(start_count, start_count + 41, '编码3')
        sht.cell(start_count, start_count + 42, '名称3')
        sht.cell(start_count, start_count + 43, '核算项目4')
        sht.cell(start_count, start_count + 44, '编码4')
        sht.cell(start_count, start_count + 45, '名称4')
        sht.cell(start_count, start_count + 46, '核算项目5')
        sht.cell(start_count, start_count + 47, '编码5')
        sht.cell(start_count, start_count + 48, '名称5')
        sht.cell(start_count, start_count + 49, '核算项目6')
        sht.cell(start_count, start_count + 50, '编码6')
        sht.cell(start_count, start_count + 51, '名称6')
        sht.cell(start_count, start_count + 52, '核算项目7')
        sht.cell(start_count, start_count + 53, '编码7')
        sht.cell(start_count, start_count + 54, '名称7')
        sht.cell(start_count, start_count + 55, '核算项目8')
        sht.cell(start_count, start_count + 56, '编码8')
        sht.cell(start_count, start_count + 57, '名称8')
        sht.cell(start_count, start_count + 58, '发票号')
        sht.cell(start_count, start_count + 59, '换票证号')
        sht.cell(start_count, start_count + 60, '客户')
        sht.cell(start_count, start_count + 61, '费用类别')
        sht.cell(start_count, start_count + 62, '收款人')
        sht.cell(start_count, start_count + 63, '物料')
        sht.cell(start_count, start_count + 64, '财务组织')
        sht.cell(start_count, start_count + 65, '供应商')
        sht.cell(start_count, start_count + 66, '辅助账业务日期')
        sht.cell(start_count, start_count + 67, '到期日')
        row = 1
        for voucher in voucher_list:
            booked_date = datetime.date(*(list(map(int, voucher.get('bookedDate').split('-')))))
            biz_date = datetime.date(*(list(map(int, voucher.get('bizDate').split('-')))))
            sht.cell(row + start_count, 0 + start_count, voucher.get('companyNumber'))
            sht.cell(row + start_count, 1 + start_count, booked_date)
            sht.cell(row + start_count, 2 + start_count, biz_date)
            sht.cell(row + start_count, 3 + start_count, voucher.get('periodNumber'))
            sht.cell(row + start_count, 4 + start_count, voucher.get('voucherType'))
            sht.cell(row + start_count, 5 + start_count, voucher.get('voucherNumber'))
            sht.cell(row + start_count, 6 + start_count, voucher.get('entrySeq'))
            sht.cell(row + start_count, 7 + start_count, voucher.get('voucherAbstract'))
            sht.cell(row + start_count, 8 + start_count, voucher.get('accountNumber'))
            sht.cell(row + start_count, 10 + start_count, voucher.get('currencyNumber'))
            sht.cell(row + start_count, 11 + start_count, voucher.get('localRate'))
            entry_DC = '0' if voucher.get('entryDC') == -1 else '1'
            sht.cell(row + start_count, 12 + start_count, entry_DC)
            sht.cell(row + start_count, 13 + start_count, voucher.get('originalAmount'))
            sht.cell(row + start_count, 14 + start_count, 0)
            sht.cell(row + start_count, 15 + start_count, 0)
            sht.cell(row + start_count, 16 + start_count, voucher.get('debitAmount'))
            sht.cell(row + start_count, 17 + start_count, voucher.get('creditAmount'))
            sht.cell(row + start_count, 18 + start_count, voucher.get('creator'))
            sht.cell(row + start_count, 21 + start_count, 0)
            sht.cell(row + start_count, 22 + start_count, 'FALSE')
            if voucher.get('asstActType1') is not None and voucher.get('asstActType1') != 'None':
                sht.cell(row + start_count, 33 + start_count, voucher.get('voucherAbstract'))
                sht.cell(row + start_count, 34 + start_count, voucher.get('asstActType1'))
                sht.cell(row + start_count, 35 + start_count, voucher.get('asstActNumber1'))
            row += 1
    else:
        start_count = 0
        wb = xlwt.Workbook()
        sht = wb.add_sheet('凭证')
        sht_cf = wb.add_sheet('现金流量')
        sht.write(start_count, start_count + 0, '公司')
        sht.write(start_count, start_count + 1, '记账日期')
        sht.write(start_count, start_count + 2, '业务日期')
        sht.write(start_count, start_count + 3, '会计期间')
        sht.write(start_count, start_count + 4, '凭证类型')
        sht.write(start_count, start_count + 5, '凭证号')
        sht.write(start_count, start_count + 6, '分录号')
        sht.write(start_count, start_count + 7, '摘要')
        sht.write(start_count, start_count + 8, '科目')
        sht.write(start_count, start_count + 9, '科目名称')
        sht.write(start_count, start_count + 10, '币种')
        sht.write(start_count, start_count + 11, '汇率')
        sht.write(start_count, start_count + 12, '方向')
        sht.write(start_count, start_count + 13, '原币金额')
        sht.write(start_count, start_count + 14, '数量')
        sht.write(start_count, start_count + 15, '单价')
        sht.write(start_count, start_count + 16, '借方金额')
        sht.write(start_count, start_count + 17, '贷方金额')
        sht.write(start_count, start_count + 18, '制单人')
        sht.write(start_count, start_count + 19, '过账人')
        sht.write(start_count, start_count + 20, '审核人')
        sht.write(start_count, start_count + 21, '附件数量')
        sht.write(start_count, start_count + 22, '过账标记')
        sht.write(start_count, start_count + 23, '机制凭证模块')
        sht.write(start_count, start_count + 24, '删除标记')
        sht.write(start_count, start_count + 25, '凭证序号')
        sht.write(start_count, start_count + 26, '单位')
        sht.write(start_count, start_count + 27, '参考信息')
        sht.write(start_count, start_count + 28, '是否有现金流量')
        sht.write(start_count, start_count + 29, '现金流量标记')
        sht.write(start_count, start_count + 30, '业务编号')
        sht.write(start_count, start_count + 31, '结算方式')
        sht.write(start_count, start_count + 32, '结算号')
        sht.write(start_count, start_count + 33, '辅助账摘要')
        sht.write(start_count, start_count + 34, '核算项目1')
        sht.write(start_count, start_count + 35, '编码1')
        sht.write(start_count, start_count + 36, '名称1')
        sht.write(start_count, start_count + 37, '核算项目2')
        sht.write(start_count, start_count + 38, '编码2')
        sht.write(start_count, start_count + 39, '名称2')
        sht.write(start_count, start_count + 40, '核算项目3')
        sht.write(start_count, start_count + 41, '编码3')
        sht.write(start_count, start_count + 42, '名称3')
        sht.write(start_count, start_count + 43, '核算项目4')
        sht.write(start_count, start_count + 44, '编码4')
        sht.write(start_count, start_count + 45, '名称4')
        sht.write(start_count, start_count + 46, '核算项目5')
        sht.write(start_count, start_count + 47, '编码5')
        sht.write(start_count, start_count + 48, '名称5')
        sht.write(start_count, start_count + 49, '核算项目6')
        sht.write(start_count, start_count + 50, '编码6')
        sht.write(start_count, start_count + 51, '名称6')
        sht.write(start_count, start_count + 52, '核算项目7')
        sht.write(start_count, start_count + 53, '编码7')
        sht.write(start_count, start_count + 54, '名称7')
        sht.write(start_count, start_count + 55, '核算项目8')
        sht.write(start_count, start_count + 56, '编码8')
        sht.write(start_count, start_count + 57, '名称8')
        sht.write(start_count, start_count + 58, '发票号')
        sht.write(start_count, start_count + 59, '换票证号')
        sht.write(start_count, start_count + 60, '客户')
        sht.write(start_count, start_count + 61, '费用类别')
        sht.write(start_count, start_count + 62, '收款人')
        sht.write(start_count, start_count + 63, '物料')
        sht.write(start_count, start_count + 64, '财务组织')
        sht.write(start_count, start_count + 65, '供应商')
        sht.write(start_count, start_count + 66, '辅助账业务日期')
        sht.write(start_count, start_count + 67, '到期日')
        row = 1
        for voucher in voucher_list:
            booked_date = datetime.date(*(list(map(int, voucher.get('bookedDate').split('-')))))
            biz_date = datetime.date(*(list(map(int, voucher.get('bizDate').split('-')))))
            sht.write(row + start_count, 0 + start_count, voucher.get('companyNumber'))
            sht.write(row + start_count, 1 + start_count, booked_date)
            sht.write(row + start_count, 2 + start_count, biz_date)
            sht.write(row + start_count, 3 + start_count, voucher.get('periodNumber'))
            sht.write(row + start_count, 4 + start_count, voucher.get('voucherType'))
            sht.write(row + start_count, 5 + start_count, voucher.get('voucherNumber'))
            sht.write(row + start_count, 6 + start_count, voucher.get('entrySeq'))
            sht.write(row + start_count, 7 + start_count, voucher.get('voucherAbstract'))
            sht.write(row + start_count, 8 + start_count, voucher.get('accountNumber'))
            sht.write(row + start_count, 10 + start_count, voucher.get('currencyNumber'))
            sht.write(row + start_count, 11 + start_count, voucher.get('localRate'))
            entry_DC = '0' if voucher.get('entryDC') == -1 else '1'
            sht.write(row + start_count, 12 + start_count, entry_DC)
            sht.write(row + start_count, 13 + start_count, voucher.get('originalAmount'))
            sht.write(row + start_count, 14 + start_count, 0)
            sht.write(row + start_count, 15 + start_count, 0)
            sht.write(row + start_count, 16 + start_count, voucher.get('debitAmount'))
            sht.write(row + start_count, 17 + start_count, voucher.get('creditAmount'))
            sht.write(row + start_count, 18 + start_count, voucher.get('creator'))
            sht.write(row + start_count, 21 + start_count, 0)
            sht.write(row + start_count, 22 + start_count, 'FALSE')
            if voucher.get('asstActType1') is not None and voucher.get('asstActType1') != 'None':
                sht.write(row + start_count, 33 + start_count, voucher.get('voucherAbstract'))
                sht.write(row + start_count, 34 + start_count, voucher.get('asstActType1'))
                sht.write(row + start_count, 35 + start_count, voucher.get('asstActNumber1'))
            row += 1
    if module == 'openpyxl':
        tail = '.xlsx'
    else:
        tail = '.xls'
    del wb['Sheet']
    wb.save(file_name_without_tail + tail)
    return file_name_without_tail + tail


def login_service(user_name='dwwb', user_pwd=None) -> str:
    global Domain
    if user_name == 'dwwb':
        user_pwd = '123456'
    try:
        userName = user_name
        password = user_pwd
        slnName = 'eas'
        dcName = 'dw2011'
        language = 'L2'
        dbType = 0
        authPattern = 'BaseDB'
        client = Client(Domain + '/ormrpc/services/EASLogin?wsdl')
        result = client.service.login(userName, password, slnName, dcName, language, dbType, authPattern)
        ssid = re.search('sessionId = "' + '(.*?)' + '"', str(result)).group(1)
        print('webservice login successfully， login as ' + userName)
        return ssid
    except Exception as e:
        manual_error_log(str(e))
        print(e)
        return 'Login Error - error message: ' + str(e)


@function_log
def import_voucher_by_id(voucher_id: str, ssid: str, monitoring=True) -> dict:
    """

    :param voucher_id:
    :param ssid:
    :param monitoring:
    :return:
    """
    voucher_session = create_session()
    voucher_record_list = voucher_session.query(tEntry).filter(tEntry.r_voucher_id == voucher_id).all()
    voucher_list = []
    for voucher in voucher_record_list:
        voucher_line = {
            'companyNumber': voucher.r_companyNumber,
            'bookedDate': voucher.r_bookedDate,
            'bizDate': voucher.r_bizDate,
            'periodYear': voucher.r_periodYear,
            'periodNumber': voucher.r_periodNumber,
            'voucherType': voucher.r_voucherType,
            'voucherNumber': voucher.r_voucherNumber,
            'entrySeq': voucher.r_entrySeq,
            'voucherAbstract': voucher.r_voucherAbstract,
            'accountNumber': voucher.r_accountNumber,
            'currencyNumber': voucher.r_currencyNumber,
            'localRate': voucher.r_localRate,
            'entryDC': voucher.r_entryDC,
            'originalAmount': voucher.r_originalAmount,
            'debitAmount': voucher.r_debitAmount,
            'creditAmount': voucher.r_creditAmount,
            'creator': voucher.r_creator,
        }
        if voucher.r_asstActType1:
            voucher_line['asstActType1'] = voucher.r_asstActType1
            voucher_line['asstActNumber1'] = voucher.r_asstActNumber1
        if voucher.r_asstActType2:
            voucher_line['asstActType2'] = voucher.r_asstActType2
            voucher_line['asstActNumber2'] = voucher.r_asstActNumber2
        voucher_list.append(voucher_line)
    for each in voucher_list:
        print(each)
    result_dict = _importVoucher(voucher_list=voucher_list, ssid=ssid)
    voucher_record = voucher_session.query(tVoucher).filter(tVoucher.r_id == voucher_id).first()
    workorder_id = voucher_record.r_work_order_id
    voucher_record.r_import_result = str(result_dict['origin_result_str'])
    if result_dict['result'] == '0000':
        voucher_record.r_kd_voucher_number = str(result_dict['voucher_number'])
    voucher_session.commit()
    all_workorder_voucher = voucher_session.query(tVoucher).filter(
        tVoucher.r_work_order_id == workorder_id
    ).all()
    finished = 0
    for each_record in all_workorder_voucher:
        if each_record.r_kd_voucher_number is None:
            finished = -1
    if finished == 0:
        voucher_session.query(tWorkOrderMain).filter(tWorkOrderMain.r_id == workorder_id).update({
            tWorkOrderMain.r_doc_status: '0'
        }, synchronize_session=False)
        voucher_session.commit()
    voucher_session.close()
    if monitoring:
        print(result_dict)
    return result_dict


@function_log
def import_voucher_from_local_data(voucher_list: list, ssid: str, workorder_id: str, monitoring=True) -> dict:
    """
    Coated function which could be used by user, also write to sql.
    :param voucher_list:
    :param ssid:
    :param monitoring: if print result in console terminal
    :return: return original result
    error code: -1: more than one voucher; -2: sql error;
    """
    # 检查传入凭证是否为同一张
    for voucher in voucher_list:
        if voucher['voucherNumber'] != voucher_list[0]['voucherNumber']:
            return {
                'result': '-1',
                'type': None,
                'year': None,
                'month': None,
                'invoke_msg': None,
                'voucher_number': None,
                'origin_result_str': 'input data error： input more than 1 voucher!',
            }
    import_voucher_session = create_session()
    voucher_id = save_voucher_into_database(voucher_list=voucher_list,
                                            work_order_id=workorder_id,)
    if not re.match('UID.*', voucher_id):
        return {
            'result': '-2',
            'type': None,
            'year': None,
            'month': None,
            'invoke_msg': None,
            'voucher_number': None,
            'origin_result_str': 'voucher information update error!',
        }
    # try to import voucher:
    result_dict = _importVoucher(voucher_list=voucher_list, ssid=ssid)
    voucher_record = import_voucher_session.query(tVoucher).filter(tVoucher.r_id == voucher_id).first()
    voucher_record.r_import_result = result_dict['origin_result_str']
    if result_dict['result'] == '0000':
        voucher_record.r_kd_voucher_number = result_dict['voucher_number']
    import_voucher_session.commit()
    import_voucher_session.close()
    if monitoring:
        print(result_dict)
    return result_dict


@function_log
def delete_voucher(entity_code:str, period: str, voucher_number:str, desc: str, ssid: str) -> int:
    global Domain
    # 开始引入
    try:
        client = Client(Domain + '/ormrpc/services/WSGLWebServiceFacade?wsdl')
        client.set_options(timeout=300)
        ssn = Element('SessionId').setText(ssid).setPrefix(p='ns1', u='http://login.webservice.bos.kingdee.com')
        client.set_options(soapheaders=ssn)
        result = client.service.deleteVoucher(entity_code, period, voucher_number, desc)
    except Exception as e:
        manual_error_log(str(e))
        result = -1
        print(e)
    return int(result)


def import_vouchers_from_local_data(voucher_list: list, ssid: str, workorder_id: str, renumber=False,
                                    monitoring=True) -> list:
    voucher_dict = _seperate_voucher_list(voucher_list=voucher_list, renumber=renumber)
    result_list = []
    for voucher_number in voucher_dict:
        curr_result = import_voucher_from_local_data(voucher_dict[voucher_number], ssid, workorder_id, monitoring)
        result_list.append(curr_result)
    return result_list


def get_account_balance(entity_code: str, str_year: str, str_month: str, ssid: str) -> list:
    global Domain
    result = []
    try:
        client = Client(Domain + '/ormrpc/services/WSGLWebServiceFacade?wsdl')
        client.set_options(timeout=300)
        ssn = Element('SessionId').setText(ssid).setPrefix(p='ns1', u='http://login.webservice.bos.kingdee.com')
        client.set_options(soapheaders=ssn)
        len_of_tb = 1001
        loop = 0
        result = []
        while len_of_tb == 1001:
            curr_result = client.service.getAccountBalance(entity_code, str(str_year), str(str_month),
                                                           loop * len_of_tb, loop * len_of_tb + len_of_tb)
            result += curr_result if curr_result is not None else []
            len_of_tb = len(curr_result) if curr_result is not None else 0
            loop += 1
        print('Invoke getAccountBalance interface successfully, param: entity_code' + entity_code + '; period: '
              + str(str_year) + '-' + str(str_month) + '; data length: ' + str(len(result) if result else 0))
    except ConnectionResetError:
        get_account_balance(entity_code, str_year, str_month, ssid)
    except Exception as e:
        result = [e]
        print('Unexpected error occurred in invoking getAccountBalance interface, param: ' + entity_code + '; period: '
              + str(str_year) + '-' + str(str_month))
        print(e)
    finally:
        return result


def get_assit_balance(entity_code: str, account_number: str, str_year: str, str_month: str, ssid: str) -> list:
    global Domain
    result = []
    try:
        client = Client(Domain + '/ormrpc/services/WSGLWebServiceFacade?wsdl')
        client.set_options(timeout=300)
        ssn = Element('SessionId').setText(ssid).setPrefix(p='ns1', u='http://login.webservice.bos.kingdee.com')
        client.set_options(soapheaders=ssn)
        # no loop designed, assit ledger is not supposed to be longer than 1k
        result = client.service.getAssitBalance(entity_code, account_number, str(str_year), str(str_month), 0, 0)
        print('Invoke getAsstBalance interface successfully, param: entity_code:' + entity_code +
              ';account_code:' + str(account_number) + '; period: '
              + str(str_year) + '-' + str(str_month) + '; data length: ' + str(len(result) if result else 0))
    except ConnectionResetError:
        get_assit_balance(entity_code, account_number, str_year, str_month, ssid)
    except Exception as e:
        result = [e]
        print('Unexpected error occurred in invoking getAsstBalance interface, param: ' + entity_code +
              ';account_code: ' + str(account_number) + '; period: '
              + str(str_year) + '-' + str(str_month))
        print(e)
    finally:
        return result


def get_account_asst_type(entity_code: str, account_number: str) -> list:
    """

    :param entity_code:
    :param account_number:
    :return: empty list implies args not exist, ['None'] implies no asst ledger
    """
    asst_type_list = []
    asst_session = create_session()
    entity_record = asst_session.query(tBaseKdEntity).filter(tBaseKdEntity.r_code == entity_code).first()
    if not entity_record:
        return asst_type_list
    account_record = asst_session.query(tBaseKdAccount).filter(
        tBaseKdAccount.r_account_code == account_number,
        tBaseKdAccount.r_entity_id == entity_record.r_id
    ).first()
    if not account_record:
        return asst_type_list
    else:
        return str(account_record.r_asst_type).replace('{', '').replace('}', '').split(',')


def get_entity_chs_name_dict() -> dict:
    chs_session = create_session()
    chs_record = chs_session.query(tBaseKdEntity.r_code, tBaseKdEntity.r_chs_name).all()
    result_dict = {}
    for record in chs_record:
        result_dict[record[0]] = record[1]
    chs_session.close()
    return result_dict


def get_default_supplier_dict() -> dict:
    s_session = create_session()
    s_record = s_session.query(tBaseKdEntity.r_code, tBaseKdSupplier.r_code).filter(
        tBaseKdSupplier.r_entity_id == tBaseKdEntity.r_id).all()
    result_dict = {}
    for record in s_record:
        result_dict[record[0]] = record[1]
    s_session.close()
    return result_dict


def get_default_cost_center_dict() -> dict:
    c_session = create_session()
    c_record = c_session.query(tBaseKdEntity.r_code, tBaseKdCostCenter.r_code).filter(
        tBaseKdCostCenter.r_entity_id == tBaseKdEntity.r_id).all()
    result_dict = {}
    for record in c_record:
        result_dict[record[0]] = record[1]
    c_session.close()
    return result_dict


def get_account_full_name_dict(entity_code=None, entity_id=None) -> dict:
    if entity_code is None and entity_id is None:
        raise Exception('function get_account_full_name_dict error occurred: need at least one arg')
    else:
        if 'UID' in entity_code:
            entity_id = entity_code
            entity_code = None
        account_session = create_session()
        if entity_code:
            entity_id = account_session.query(tBaseKdEntity.r_id).filter(
                tBaseKdEntity.r_code == entity_code
            ).first()[0]
        account_dict = {}
        account_record = account_session.query(tBaseKdAccount.r_account_code,
                                               tBaseKdAccount.r_account_full_name).filter(
            tBaseKdAccount.r_entity_id == entity_id
        ).all()
        for each in account_record:
            account_dict.setdefault(each.r_account_code, each.r_account_full_name)
        return account_dict


def save_voucher_into_database(voucher_list: list, work_order_id: str) -> str:
    """
    one voucher at a time
    :param voucher_list:
    :param work_order_id:
    :param description:
    :return:
    'UID.*': OK
    '-1': sql update error
    '-2': more than one voucher input
    """
    if voucher_list:
        for voucher in voucher_list:
            if voucher['voucherNumber'] != voucher_list[0]['voucherNumber']:
                return '-2'
        total_debit = 0
        total_credit = 0
        for voucher in voucher_list:
            total_debit += float(voucher['debitAmount'])
            total_credit += float(voucher['creditAmount'])
        voucher_info_session = create_session()
        voucher_in_db = voucher_info_session.query(tVoucher).filter(
            tVoucher.r_sys_voucher_number == voucher_list[0]['voucherNumber'],
            tVoucher.r_work_order_id == work_order_id
        ).first()
        try:
            if voucher_in_db:
                voucher_id = voucher_in_db.r_id
                voucher_info_session.query(tVoucher).filter(tVoucher.r_id == voucher_id).update({
                    tVoucher.r_sys_voucher_number: str(voucher_list[0]['voucherNumber']),
                    tVoucher.r_entity_code: str(voucher_list[0].get('companyNumber')),
                    tVoucher.r_abstract: str(voucher_list[0].get('voucherAbstract')),
                    tVoucher.r_year: int(voucher_list[0].get('periodYear')),
                    tVoucher.r_month: int(voucher_list[0].get('periodMonth')),
                    tVoucher.r_type: str(voucher_list[0].get('voucherType')),
                    tVoucher.r_credit_total: float(total_credit),
                    tVoucher.r_debit_total: float(total_debit),
                    tVoucher.r_description: str(voucher_list[0].get('description')),
                }, synchronize_session=False)
                voucher_info_session.commit()
                voucher_info_session.query(tEntry).filter(tEntry.r_voucher_id == voucher_id).delete(synchronize_session=False)
                voucher_info_session.commit()
            else:
                voucher_id = generate_key()
                voucher_info_session.add(tVoucher(
                    r_id=str(voucher_id),
                    r_work_order_id=str(work_order_id),
                    r_sys_voucher_number=str(voucher_list[0].get('voucherNumber')),
                    r_entity_code=str(voucher_list[0].get('companyNumber')),
                    r_abstract=str(voucher_list[0].get('voucherAbstract')),
                    r_year=int(voucher_list[0].get('periodYear')),
                    r_month=int(voucher_list[0].get('periodNumber')),
                    r_type=str(voucher_list[0].get('voucherType')),
                    r_debit_total=float(total_debit),
                    r_credit_total=float(total_credit),
                    r_description=str(voucher_list[0].get('description'))
                ))
                voucher_info_session.commit()
            for voucher in voucher_list:
                voucher_info_session.add(tEntry(
                    r_id=str(generate_key()),
                    r_voucher_id=str(voucher_id),
                    r_companyNumber=str(voucher.get('companyNumber')),
                    r_bookedDate=str(voucher.get('bookedDate')),
                    r_bizDate=str(voucher.get('bizDate')),
                    r_periodYear=int(voucher.get('periodYear')),
                    r_periodNumber=int(voucher.get('periodNumber')),
                    r_voucherType=str(voucher.get('voucherType')),
                    r_voucherNumber=str(voucher.get('voucherNumber')),
                    r_entrySeq=int(voucher.get('entrySeq')),
                    r_voucherAbstract=str(voucher.get('voucherAbstract')),
                    r_accountNumber=str(voucher.get('accountNumber')),
                    r_currencyNumber=str(voucher.get('currencyNumber')),
                    r_localRate=float(voucher.get('localRate')) if voucher.get('localRate') else 1,
                    r_entryDC=int(voucher.get('entryDC')),
                    r_originalAmount=float(voucher.get('originalAmount')),
                    r_debitAmount=float(voucher.get('debitAmount')),
                    r_creditAmount=float(voucher.get('creditAmount')),
                    r_creator=str(voucher.get('creator')),
                    r_description=str(voucher.get('description')),
                    r_asstActType1=str(voucher.get('asstActType1')),
                    r_asstActNumber1=str(voucher.get('asstActNumber1')),
                    r_asstActType2=str(voucher.get('asstActType2')),
                    r_asstActNumber2=str(voucher.get('asstActNumber2')),
                ))
                voucher_info_session.commit()
            voucher_info_session.close()
            return voucher_id
        except Exception as e:
            manual_error_log(str(e))
            print(e)
            voucher_info_session.rollback()
            return '-1'
    else:
        return '-2'


def save_vouchers_into_database(voucher_list: list, work_order_id: str, description: [None, str],
                                renumber=False) -> str:
    separated_voucher_dict = _seperate_voucher_list(voucher_list, renumber=renumber)
    all_success = '0'
    for voucher_number in separated_voucher_dict:
        curr_result = save_voucher_into_database(separated_voucher_dict[voucher_number], work_order_id, description)
        if len(curr_result) > 2:
            all_success = '-1'
    return all_success


def get_voucher_from_database(voucher_id: str) -> list:
    voucher_session = create_session()
    voucher_list_record = voucher_session.query(tEntry).filter(tEntry.r_voucher_id == voucher_id).all()
    result_voucher_list = []
    for voucher in voucher_list_record:
        if not voucher.r_asstActType1 and not voucher.r_asstActType2:
            result_voucher_list.append({
                'companyNumber': str(voucher.r_companyNumber),
                'bookedDate': str(voucher.r_bookedDate),
                'bizDate': str(voucher.r_bizDate),
                'periodYear': int(voucher.r_periodYear),
                'periodNumber': int(voucher.r_periodNumber),
                'voucherType': str(voucher.r_voucherType),
                'voucherNumber': str(voucher.r_voucherNumber),
                'entrySeq': int(voucher.r_entrySeq),
                'voucherAbstract': str(voucher.r_voucherAbstract),
                'accountNumber': str(voucher.r_accountNumber),
                'currencyNumber': str(voucher.r_currencyNumber),
                'localRate': float(voucher.r_localRate),
                'entryDC': int(voucher.r_entryDC),
                'originalAmount': float(voucher.r_originalAmount),
                'debitAmount': float(voucher.r_debitAmount),
                'creditAmount': float(voucher.r_creditAmount),
                'creator': str(voucher.r_creator),
            })
        elif not voucher.r_asstActType2:
            result_voucher_list.append({
                'companyNumber': str(voucher.r_companyNumber),
                'bookedDate': str(voucher.r_bookedDate),
                'bizDate': str(voucher.r_bizDate),
                'periodYear': int(voucher.r_periodYear),
                'periodNumber': int(voucher.r_periodNumber),
                'voucherType': str(voucher.r_voucherType),
                'voucherNumber': str(voucher.r_voucherNumber),
                'entrySeq': int(voucher.r_entrySeq),
                'voucherAbstract': str(voucher.r_voucherAbstract),
                'accountNumber': str(voucher.r_accountNumber),
                'currencyNumber': str(voucher.r_currencyNumber),
                'localRate': float(voucher.r_localRate),
                'entryDC': int(voucher.r_entryDC),
                'originalAmount': float(voucher.r_originalAmount),
                'debitAmount': float(voucher.r_debitAmount),
                'creditAmount': float(voucher.r_creditAmount),
                'creator': str(voucher.r_creator),
                'asstActType1': str(voucher.r_asstActType1),
                'asstActNumber1': str(voucher.r_asstActNumber1),
            })
        else:
            result_voucher_list.append({
                'companyNumber': str(voucher.r_companyNumber),
                'bookedDate': str(voucher.r_bookedDate),
                'bizDate': str(voucher.r_bizDate),
                'periodYear': int(voucher.r_periodYear),
                'periodNumber': int(voucher.r_periodNumber),
                'voucherType': str(voucher.r_voucherType),
                'voucherNumber': str(voucher.r_voucherNumber),
                'entrySeq': int(voucher.r_entrySeq),
                'voucherAbstract': str(voucher.r_voucherAbstract),
                'accountNumber': str(voucher.r_accountNumber),
                'currencyNumber': str(voucher.r_currencyNumber),
                'localRate': float(voucher.r_localRate),
                'entryDC': int(voucher.r_entryDC),
                'originalAmount': float(voucher.r_originalAmount),
                'debitAmount': float(voucher.r_debitAmount),
                'creditAmount': float(voucher.r_creditAmount),
                'creator': str(voucher.r_creator),
                'asstActType1': str(voucher.r_asstActType1),
                'asstActNumber1': str(voucher.r_asstActNumber1),
                'asstActType2': str(voucher.r_asstActType2),
                'asstActNumber2': str(voucher.r_asstActNumber2),
            })
    return result_voucher_list


if __name__ == '__main__':
    pass